import pandas as pd
from openpyxl import load_workbook

file_path = "test.xlsx"
wb = load_workbook(file_path)
contato_sheet = wb["Contato Evento Tratado"]
siscad_df = pd.read_excel(file_path, sheet_name="Base SISCAD")

# Ensure CRMV and UF are strings for consistent comparison
siscad_df["CRMV"] = siscad_df["CRMV"].astype(str).str.strip()
siscad_df["UF"] = siscad_df["UF"].astype(str).str.strip().str.upper()

def has_two_matching_words(name1, name2):
    words1 = set(str(name1).lower().split())
    words2 = set(str(name2).lower().split())
    return len(words1.intersection(words2)) >= 2

for row in range(2, contato_sheet.max_row + 1):
    name_d = contato_sheet[f"D{row}"].value
    name_k = contato_sheet[f"K{row}"].value

    if has_two_matching_words(name_d, name_k):
        # Se os nomes forem compatíveis, atualiza diretamente
        contato_sheet[f"D{row}"] = name_k
        contato_sheet[f"K{row}"] = name_k
        contato_sheet[f"S{row}"] = "Sim"
    else:
        # Só verifica CRMV e UF se os nomes não forem compatíveis
        ncrmv = contato_sheet[f"H{row}"].value  # NCRMV
        uf_crmv = contato_sheet[f"I{row}"].value  # UF CRMV

        if ncrmv is not None and uf_crmv is not None:
            ncrmv = str(ncrmv).strip()
            uf_crmv = str(uf_crmv).strip().upper()

            # Filtra com os valores normalizados
            filtered = siscad_df[
                (siscad_df["CRMV"] == ncrmv) &
                (siscad_df["UF"] == uf_crmv)
            ]

            if not filtered.empty:
                # Atualiza com o nome da base SISCAD
                nome_siscad = filtered.iloc[0]["Nome"]
                contato_sheet[f"D{row}"] = nome_siscad
                contato_sheet[f"K{row}"] = nome_siscad
                contato_sheet[f"S{row}"] = "Sim"
            else:
                contato_sheet[f"S{row}"] = "N/A SISCAD"
        else:
            contato_sheet[f"S{row}"] = "Não é VET"


from datetime import datetime


# Get current date and time
now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H%M%S")  # Format: YYYY-MM-DD_HH-MM-SS

# Construct the filename with timestamp
filename = f"updated_file_names_{timestamp}.xlsx"

wb.save(filename)
print(f"Updated: {filename}")