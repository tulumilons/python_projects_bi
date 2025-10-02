import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

# === Config ===
file_path  = "test.xlsx"
sheet_name = "Contato Evento Tratado"
start_row  = 2
source_col = "T"
target_col = "U"
blank_if_not_11 = False
overwrite = False

def only_digits(v: object) -> str:
    return "" if v is None else re.sub(r"\D+", "", str(v))

def format_br_11(d: str) -> str:
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    elif d == "":
        return ""
    else:
        return "Formato Inválido"

# Carrega o arquivo original
wb = load_workbook(file_path)
ws = wb[sheet_name] if sheet_name else wb.active

src_idx = column_index_from_string(source_col)
tgt_idx = column_index_from_string(target_col)

# Adiciona cabeçalho se necessário
if start_row > 1:
    header = ws.cell(row=1, column=src_idx).value
    ws.cell(row=1, column=tgt_idx).value = (str(header) + " (formatado)") if header else "Telefone formatado"

# Aplica a formatação
for r in range(start_row, ws.max_row + 1):
    raw = ws.cell(row=r, column=src_idx).value
    digits = only_digits(raw)
    ws.cell(row=r, column=tgt_idx).value = "" if (blank_if_not_11 and len(digits) != 11) else format_br_11(digits)


from datetime import datetime

# Gera timestamp no formato desejado
timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

# Cria o nome do novo arquivo
filename = f"updated_file_{timestamp}.xlsx"

# Salva o workbook modificado (não crie um novo com Workbook())
wb.save(filename)

print(f"Arquivo salvo como {filename}")