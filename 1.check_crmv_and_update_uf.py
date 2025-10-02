from openpyxl import load_workbook

src_path = "test.xlsx" 

wb = load_workbook(src_path)
ws = wb.active  # ou defina uma aba específica

start_row = 2  # assumindo que a linha 1 é o cabeçalho
col_h = 8      # H
col_i = 9      # I
col_j = 10     # J

def split_letters_digits(value):
    if value is None:
        return "", ""
    s = str(value)
    alnum = [ch for ch in s if ch.isalnum()]
    letters = "".join(ch for ch in alnum if ch.isalpha())
    digits  = "".join(ch for ch in alnum if ch.isdigit())
    return letters, digits

# max_row = ws.max_row
# for row in range(start_row, max_row + 1):
#     cell_h = ws.cell(row=row, column=col_h)
#     cell_i = ws.cell(row=row, column=col_i)
#     cell_j = ws.cell(row=row, column=col_j)

#     val_h = cell_h.value
#     val_i = cell_i.value
#     val_j = cell_j.value

#     # Ignora fórmulas
#     if isinstance(val_h, str) and val_h.startswith("="):
#         continue

#     letters_h, digits_h = split_letters_digits(val_h)

#     if val_i is not None and str(val_i).strip() != "":
#         # Se já tem valor em I, apenas limpa H
#         ws.cell(row=row, column=col_h, value=digits_h)
#     else:
#         # Tenta preencher I com letras de H
#         ws.cell(row=row, column=col_i, value=letters_h)
#         ws.cell(row=row, column=col_h, value=digits_h)

#         # Verifica se I ainda está vazio após isso
#         updated_i = ws.cell(row=row, column=col_i).value
#         if updated_i is None or str(updated_i).strip() == "":
#             # Se ainda vazio, tenta usar letras de J
#             letters_j, _ = split_letters_digits(val_j)
#             sorted_letters_j = "".join(sorted(letters_j))
#             ws.cell(row=row, column=col_i, value=sorted_letters_j)

max_row = ws.max_row
for row in range(start_row, max_row + 1):
    cell_h = ws.cell(row=row, column=col_h)
    cell_i = ws.cell(row=row, column=col_i)
    cell_j = ws.cell(row=row, column=col_j)

    val_h = cell_h.value
    val_i = cell_i.value
    val_j = cell_j.value

    # Ignora fórmulas
    if isinstance(val_j, str) and val_j.startswith("="):
        continue

    # Só processa se H ou I estiverem vazias
    if (val_h is None or str(val_h).strip() == "") or (val_i is None or str(val_i).strip() == ""):
        letters_j, digits_j = split_letters_digits(val_j)

        # Evita colocar "000000" se não houver dígitos
        digits_j = digits_j if digits_j != "" else None
        letters_j_sorted = "".join(sorted(letters_j)) if letters_j != "" else None

        # Atualiza apenas se estiverem vazias
        if val_h is None or str(val_h).strip() == "":
            ws.cell(row=row, column=col_h, value=digits_j)

        if val_i is None or str(val_i).strip() == "":
            ws.cell(row=row, column=col_i, value=letters_j_sorted)

            
from datetime import datetime
# Get current date and time
now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H%M%S")  # Format: YYYY-MM-DD_HH-MM-SS

# Construct the filename with timestamp
filename = f"updated_file_crmv_{timestamp}.xlsx"

wb.save(filename)
print(f"Done. Saved to {filename}")