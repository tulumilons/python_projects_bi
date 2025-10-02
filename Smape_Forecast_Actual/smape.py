import pandas as pd # Import necessary libraries
from datetime import datetime
import unicodedata
import math
from openpyxl import load_workbook

# Function to normalize text (remove accents, convert to uppercase, strip whitespace)
def normalize(text):
    if text is None or (isinstance(text, float) and math.isnan(text)):
        return ""
    return unicodedata.normalize('NFKD', str(text)).encode('ASCII', 'ignore').decode().upper().strip()

# List of Excel sheet names to be processed
sheets = [
    "Suínos",
    "AVES",
    "RUM",
    "pet-bio (goldschmidt)",
    "pet-para (pcastro)",
    "pet-para (boliveira)",
    "Equinos"
]

# ----------------------------
# 1) Build SMAPE dictionary
# ----------------------------
smape_dict = {} # Dictionary to store SMAPE values by SKU
# Loop through each sheet
for sheet in sheets:
    print(f"\n📄 Processando aba: {sheet}")
    try:
        # Read the sheet from the Excel file (no header)
        file1 = pd.read_excel(
            r"C:\Users\lgabriel\Downloads\TodasAsClassesTemplate SMAPE & TS Calculation_17_07_2025.xlsx",
            sheet_name=sheet,
            header=None
        )
    except Exception as e:
        # Handle error if sheet can't be loaded
        print(f"❌ Erro ao carregar aba '{sheet}': {e}")
        continue

    # Process data in blocks of 16 rows
    for block_start in range(1, len(file1), 16):
        block_end = block_start + 16
        block = file1.iloc[block_start:block_end]
        # Look for the row containing the label 'R3M@%SMAPE'
        for i in range(block.shape[0]):
            label = normalize(block.iloc[i, 1])
            if label == 'R3M@%SMAPE':
                sku = normalize(block.iloc[0, 0]) # SKU is in the first row of the block, first column
                smape = block.iloc[i, 10] # SMAPE value is in the same row as the label, column 11
                try:
                    smape = float(smape) # Convert to float
                    if sku:
                        smape_dict[sku] = smape # Store in dictionary
                        print(f"✅ {sheet}: {sku} → {smape:.2f}")
                except Exception as e:
                    print(f"❌ Erro ao processar SMAPE para SKU {sku}: {e}")
                break # Exit loop after finding the label

# ----------------------------
# 2) Load original formatted Excel file
# ----------------------------
wb = load_workbook(r"C:\Users\lgabriel\Downloads\ACOMPANHAMENTO FCST PERFORMANCE - Setembro.xlsx") # Load Excel file with formatting
ws = wb.active # Select active sheet

# ----------------------------
# 3) Update column X with SMAPE values
# ----------------------------
updated_rows = 0 # Counter for updated rows
missing_skus = set() # Set to store SKUs not found

# Loop through all rows in the worksheet
for idx in range(1, ws.max_row + 1):
    sku_raw = ws.cell(row=idx, column=5).value  # Read SKU from column E
    sku = normalize(sku_raw)

    if not sku:
        continue # Skip if SKU is empty

    if sku in smape_dict:
        ws.cell(row=idx, column=24).value = smape_dict[sku]  # Update column X with SMAPE value
        updated_rows += 1
        # print(f"✅ Linha {idx}: SKU {sku} → Coluna X atualizada para {smape_dict[sku]:.2f}")
    else:
        missing_skus.add(sku) # Add to list of missing SKUs

# ----------------------------
# 4) Save updated file with timestamp
# ----------------------------
timestamp = datetime.now().strftime("%Y%m%d_%H%M") # Generate timestamp for filename
output_path = fr"C:\Users\lgabriel\Downloads\final_data_updated_{timestamp}.xlsx"
wb.save(output_path) # Save updated Excel file

# Display summary of the operation
print(f"\n🎯 Atualização concluída. {updated_rows} linhas foram atualizadas.")
if missing_skus:
    print(f"⚠️ {len(missing_skus)} SKUs não encontrados nas abas de SMAPE:")
    for s in sorted(missing_skus):
        print(f" - {s}")
